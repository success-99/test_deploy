# from django.core.mail import send_mail
# from django.conf import settings
#
# def send_registration_notification_email(classes_count):
#     subject = 'New Student Registration'
#     message = f"Platformangizdan umumiy {classes_count}ta student ro'yxatdan o'tdi."
#     from_email = settings.DEFAULT_FROM_EMAIL
#     recipient_list = [settings.ADMIN_EMAIL]  # Administrator email
#     send_mail(subject, message, from_email, recipient_list)
# import os
# from django.conf import settings
#
# os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'onlinequiz.settings')
# settings.configure()
# # Faylni o'chirish uchun kerakli yo'l
# file_to_delete = os.path.join(settings.MEDIA_ROOT, 'car.jpg')
#
# # Faylni o'chirish
# if os.path.exists(file_to_delete):
#     os.remove(file_to_delete)
#     print(f"{file_to_delete} o'chirildi.")
# else:
#     print(f"{file_to_delete} topilmadi.")

# import os
# from django.conf import settings
# os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'onlinequiz.settings')
# settings.configure()
# # Media papkasining joylashgan manzili
# media_root = settings.MEDIA_ROOT
# print(media_root)
#
# # Media papkasining ichidagi barcha fayllar va papkalarni ko'rish
# for root, dirs, files in os.walk(media_root):
#     for file in files:
#         file_path = os.path.join(root, file)
#         print(f"Fayl: {file_path}")
#     for directory in dirs:
#         dir_path = os.path.join(root, directory)
#         print(f"Papka: {dir_path}")
# import os
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
#
# # Media papkasining manzili
# media_root = os.path.join(BASE_DIR, 'staticfile')
#
# # Media papkasining ichidagi fayllar va papkalarni ko'rish
# for item in os.listdir(media_root):
#     item_path = os.path.join(media_root, item)
#     if os.path.isfile(item_path):
#         print(f"Fayl: {item}")
#     elif os.path.isdir(item_path):
#         print(f"Papka: {item}")

# import os
#
# # Media papkasining manzili
# media_root = '/Users/mbp13/Desktop/3-kurs 1-smestr/online-test-1/media'  # Media papkasining to'g'ri manzilini o'zgartiring
#
# # Media papkasining ichidagi fayllar va papkalarni ko'rish
# for item in os.listdir(media_root):
#     item_path = os.path.join(media_root, item)
#     if os.path.isfile(item_path):
#         print(f"Fayl: {item}")
#     elif os.path.isdir(item_path):
#         print(f"Papka: {item}")

# from django.core.files.storage import default_storage
# from django.conf import settings
import os
import django
#
# def delete_uploaded_images(upload_dir):
#     os.environ.setdefault("DJANGO_SETTINGS_MODULE", "onlinequiz.settings")
#     django.setup()
#     media_root = settings.MEDIA_ROOT
#     print(media_root)
#     image = []
#
#     # Get a list of uploaded image files
#     for root, dirs, files in os.walk(os.path.join(media_root, upload_dir)):
#         for file in files:
#             if file.lower().endswith('.png', '.jpg'):
#                 image.append(os.path.join(root, file))
#
#     # Delete the image files
#     for image in image:
#         default_storage.delete(image)
#
# # Bu funksiyani qo'shing va upload_dir o'zgaruvchisini uzating
# upload_dir = 'uploads/'  # O'zgartiring kerak bo'lgan yo'liga
# delete_uploaded_images(upload_dir)

# delete_uploaded_images(settings.CKEDITOR_UPLOAD_PATH)


# from django.core.files.storage import default_storage
# from django.conf import settings2
#
# def delete_uploaded_images():
#     os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'onlinequiz.settings')
#     settings.configure()
#     media_root = settings.MEDIA_ROOT
#     upload_dir = settings.CKEDITOR_UPLOAD_PATH
#     image = []
#
#     # Get a list of uploaded image files
#     for root, dirs, files in os.walk(os.path.join(media_root, upload_dir)):
#         for file in files:
#             if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.svg')):
#                 image.append(os.path.join(root, file))
#
#     # Delete the image files
#     for image in image:
#         default_storage.delete(image)
#
# # Call the function to delete uploaded image
# delete_uploaded_images()
# import pandas as pd
# from openpyxl import Workbook
#
# # Test topshirgan studentlar va natijalari
# students_results = [
#     {'student_name': 'John', 'subject': 'Math', 'score': 90},
#     {'student_name': 'Alice', 'subject': 'Math', 'score': 85},
#     {'student_name': 'Bob', 'subject': 'Math', 'score': 92},
#     # Yana studentlar va natijalari
# ]
#
# # Ma'lumotlarni DataFramega o'zlashtirish
# df = pd.DataFrame(students_results)
#
# # XLSX faylini yaratish va natijalarni yozish
# wb = Workbook()
# ws = wb.active
#
# # Jadvallarni yozish
# for r_idx, row in enumerate(df.iterrows(), start=2):
#     for c_idx, value in enumerate(row[1], start=1):
#         ws.cell(row=r_idx, column=c_idx, value=value)
#
# # Faylni saqlash
# wb.save('students_results.xlsx')


import pandas as pd
from openpyxl import Workbook
from django.conf import settings

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'onlinequiz.settings')
settings.configure()
from quiz.models import Result

# latest_date = Result.objects.latest('date').date
# print(latest_date)
# results = Result.objects.filter(date=latest_date).select_related('student', 'course', 'classes')
# print(results)

# for result in results:
#     student = result.student
#     data.append({
#         'Ism': student.user.first_name,
#         'Familiya': student.user.last_name,
#         'Kurs': result.course.course_name,
#         'Baho': result.marks,
#         'Sana': result.date.strftime('%Y-%m-%d %H:%M:%S'),
#     })

# df = pd.DataFrame(data)
# print(df)
# wb = Workbook()
# ws = wb.active

# for r_idx, row in enumerate(df.iterrows(), start=2):
#     for c_idx, value in enumerate(row[1], start=1):
#         ws.cell(row=r_idx, column=c_idx, value=value)

# wb.save('student_results_latest_date.xlsx')
